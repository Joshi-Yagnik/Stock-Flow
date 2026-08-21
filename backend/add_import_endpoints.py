import sys

file_path = r'd:\Stock-Flow\backend\app\routers\products.py'

new_code = '''
import io
import openpyxl
from openpyxl.styles import Font, PatternFill
from fastapi import UploadFile, File, Form, HTTPException, status
from fastapi.responses import StreamingResponse
from app.schemas.schemas import (
    ProductImportRow,
    ProductImportPreviewResponse,
    ProductImportExecuteRequest,
    ProductImportExecuteResponse
)
from app.models.models import Category
from decimal import Decimal, InvalidOperation

@router.get("/import/template")
async def download_import_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"

    headers = [
        "Product Name", "SKU", "Category", 
        "Selling Price", "Cost Price", "Stock", "Unit"
    ]
    
    # Write headers with styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

    # Sample row
    sample_data = ["Sample Product", "SKU-001", "Electronics", 999.99, 800.00, 50, "pcs"]
    for col, value in enumerate(sample_data, start=1):
        ws.cell(row=2, column=col, value=value)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="StockFlow_Products_Template.xlsx"'}
    )

@router.post("/import/preview", response_model=ProductImportPreviewResponse)
async def preview_excel_import(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db)
):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Invalid file format. Please upload an Excel file.")

    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read Excel file: {str(e)}")

    # Get existing SKUs
    existing_products_result = await db.execute(
        select(Product.sku).where(
            Product.owner_id == current_user.id,
            Product.sku.isnot(None)
        )
    )
    existing_skus = {sku.lower() for sku in existing_products_result.scalars().all()}
    
    # Get existing Categories
    existing_categories_result = await db.execute(
        select(Category.name).where(Category.owner_id == current_user.id)
    )
    existing_category_names = {name.lower() for name in existing_categories_result.scalars().all()}

    rows_data = []
    headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]
    
    expected_headers = ["product name", "sku", "category", "selling price", "cost price", "stock", "unit"]
    
    # Ensure minimum required columns exist
    if "product name" not in headers or "category" not in headers or "selling price" not in headers:
         raise HTTPException(status_code=400, detail="Missing required columns. Please use the template.")

    idx_name = headers.index("product name") if "product name" in headers else -1
    idx_sku = headers.index("sku") if "sku" in headers else -1
    idx_cat = headers.index("category") if "category" in headers else -1
    idx_sell = headers.index("selling price") if "selling price" in headers else -1
    idx_cost = headers.index("cost price") if "cost price" in headers else -1
    idx_stock = headers.index("stock") if "stock" in headers else -1
    idx_unit = headers.index("unit") if "unit" in headers else -1

    valid_count = 0
    invalid_count = 0
    duplicate_count = 0
    new_cats = set()

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Skip completely empty rows
        if not any(row):
            continue
            
        errors = []
        name = str(row[idx_name]).strip() if idx_name >= 0 and row[idx_name] is not None else ""
        sku = str(row[idx_sku]).strip() if idx_sku >= 0 and row[idx_sku] is not None else ""
        category = str(row[idx_cat]).strip() if idx_cat >= 0 and row[idx_cat] is not None else ""
        selling_price_raw = row[idx_sell] if idx_sell >= 0 else None
        cost_price_raw = row[idx_cost] if idx_cost >= 0 else None
        stock_raw = row[idx_stock] if idx_stock >= 0 else 0
        unit = str(row[idx_unit]).strip() if idx_unit >= 0 and row[idx_unit] is not None else "pcs"

        if not name:
            errors.append("Product Name is missing")
        if not category:
            errors.append("Category is missing")

        # Validate Selling Price
        selling_price = None
        try:
            selling_price = Decimal(str(selling_price_raw))
            if selling_price <= 0:
                errors.append("Selling price must be > 0")
        except (InvalidOperation, TypeError, ValueError):
            errors.append("Invalid Selling Price")

        # Validate Cost Price
        cost_price = None
        if cost_price_raw is not None and str(cost_price_raw).strip() != "":
            try:
                cost_price = Decimal(str(cost_price_raw))
                if cost_price < 0:
                    errors.append("Cost price cannot be negative")
            except (InvalidOperation, TypeError, ValueError):
                errors.append("Invalid Cost Price")

        # Validate Stock
        stock = 0
        try:
            stock = int(float(str(stock_raw)))
            if stock < 0:
                errors.append("Stock cannot be negative")
        except (TypeError, ValueError):
            errors.append("Invalid Stock value")

        status = "valid"
        if errors:
            status = "invalid"
            invalid_count += 1
        elif sku and sku.lower() in existing_skus:
            status = "duplicate"
            errors.append("SKU already exists")
            duplicate_count += 1
        else:
            valid_count += 1
            if sku:
                # Add to set so we catch duplicates within the same excel file
                existing_skus.add(sku.lower())

        if category and category.lower() not in existing_category_names:
            new_cats.add(category.lower())

        rows_data.append(ProductImportRow(
            row_number=row_idx,
            name=name,
            sku=sku if sku else None,
            category=category,
            selling_price=selling_price or Decimal("0"),
            cost_price=cost_price,
            stock=stock,
            unit=unit if unit else "pcs",
            status=status,
            errors=errors
        ))

    return ProductImportPreviewResponse(
        total_rows=len(rows_data),
        valid_rows=valid_count,
        invalid_rows=invalid_count,
        duplicate_rows=duplicate_count,
        new_categories=len(new_cats),
        rows=rows_data
    )

@router.post("/import/execute", response_model=ProductImportExecuteResponse)
async def execute_excel_import(
    request: ProductImportExecuteRequest,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db)
):
    valid_rows = [r for r in request.rows if r.status == "valid"]
    if not valid_rows:
        return ProductImportExecuteResponse(
            products_imported=0,
            categories_created=0,
            products_skipped=len(request.rows),
            products_failed=0
        )

    # 1. Resolve Categories
    unique_category_names = {r.category.strip() for r in valid_rows}
    
    existing_categories_result = await db.execute(
        select(Category).where(Category.owner_id == current_user.id)
    )
    existing_categories = existing_categories_result.scalars().all()
    
    # Create mapping by lowercase name
    cat_map = {c.name.lower(): c.id for c in existing_categories}
    
    categories_created = 0
    for cat_name in unique_category_names:
        lower_name = cat_name.lower()
        if lower_name not in cat_map:
            new_cat = Category(
                owner_id=current_user.id,
                name=cat_name,
                is_active=True
            )
            db.add(new_cat)
            await db.flush() # flush to get the ID
            cat_map[lower_name] = new_cat.id
            categories_created += 1

    # 2. Re-verify existing SKUs one last time
    existing_products_result = await db.execute(
        select(Product.sku).where(
            Product.owner_id == current_user.id,
            Product.sku.isnot(None)
        )
    )
    existing_skus = {sku.lower() for sku in existing_products_result.scalars().all()}
    
    # 3. Insert Products
    products_imported = 0
    products_skipped = len(request.rows) - len(valid_rows)
    products_failed = 0
    
    new_products = []
    
    for row in valid_rows:
        if row.sku and row.sku.lower() in existing_skus:
            products_skipped += 1
            continue
            
        try:
            product = Product(
                owner_id=current_user.id,
                name=row.name,
                sku=row.sku,
                category_id=cat_map[row.category.strip().lower()],
                purchase_price=row.cost_price,
                selling_price=row.selling_price,
                stock_quantity=row.stock,
                unit=row.unit,
                is_active=True
            )
            new_products.append(product)
            if row.sku:
                existing_skus.add(row.sku.lower())
            products_imported += 1
        except Exception:
            products_failed += 1

    if new_products:
        db.add_all(new_products)
        
    await db.commit()
    
    return ProductImportExecuteResponse(
        products_imported=products_imported,
        categories_created=categories_created,
        products_skipped=products_skipped,
        products_failed=products_failed
    )
'''

with open(file_path, 'a', encoding='utf-8') as f:
    f.write(new_code)

print("Endpoints added.")
